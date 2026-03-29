"""
onecell_tool/main.py
product.xls → onecell_template.xlsx 자동 변환 도구
"""
import os
import sys
import math
import configparser
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from io import StringIO

from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
# 경로 헬퍼: PyInstaller 패키징 대응
# ─────────────────────────────────────────────
def resource_path(relative: str) -> str:
    """PyInstaller 실행 시 임시 압축해제 경로, 아니면 현재 디렉터리 반환."""
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, relative)


TEMPLATE_PATH = resource_path("onecell_template.xlsx")
SETTINGS_PATH = os.path.join(os.path.dirname(os.path.abspath(sys.argv[0])), "settings.ini")


# ─────────────────────────────────────────────
# Settings 파싱
# ─────────────────────────────────────────────
def load_settings() -> configparser.ConfigParser:
    cfg = configparser.ConfigParser()
    if os.path.exists(SETTINGS_PATH):
        cfg.read(SETTINGS_PATH, encoding="utf-8")
    else:
        cfg["general"] = {"tag": ""}
        cfg["pricing"] = {"margin_rate": "15"}
    return cfg


def save_settings(cfg: configparser.ConfigParser):
    with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
        cfg.write(f)


# ─────────────────────────────────────────────
# product.xls 파싱 (HTML 위장 XLS)
# ─────────────────────────────────────────────
def parse_product_file(filepath: str) -> tuple[list[str], list[list[str]]]:
    """HTML 형식 XLS 파일을 파싱해 (헤더, 행 리스트) 반환."""
    with open(filepath, encoding="utf-8", errors="replace") as f:
        content = f.read()
    soup = BeautifulSoup(content, "html.parser")
    table = soup.find("table")
    if table is None:
        raise ValueError("파일에서 테이블을 찾을 수 없습니다.")
    rows = table.find_all("tr")
    if not rows:
        raise ValueError("테이블 행이 없습니다.")
    header = [td.get_text(strip=True) for td in rows[0].find_all(["td", "th"])]

    def cell_text(td) -> str:
        """
        <br>/<br/> 를 \n으로 변환 후 텍스트 추출.
        strip=True 는 \n 까지 제거하므로 자식 노드를 직접 순회한다.
        """
        from bs4 import NavigableString, Tag as BSTag
        parts = []
        for child in td.children:
            if isinstance(child, NavigableString):
                t = child.strip()
                if t:
                    parts.append(t)
            elif isinstance(child, BSTag) and child.name == "br":
                parts.append("\n")
        return "".join(parts).strip()

    data = []
    for tr in rows[1:]:
        cells = [cell_text(td) for td in tr.find_all(["td", "th"])]
        while len(cells) < len(header):
            cells.append("")
        data.append(cells)
    return header, data


# ─────────────────────────────────────────────
# 옵션명/옵션값 파싱
# ─────────────────────────────────────────────
OPTION_KEYWORDS = ["색상", "사이즈", "타입", "품목"]


def parse_option_names(raw_name: str) -> list[str]:
    """
    옵션명 파싱. 두 가지 형태를 모두 처리:
      1) 개행 구분: '색상\n사이즈' → ['색상', '사이즈']  (운영 파일 기준)
      2) 붙은 형태: '색상사이즈'   → 키워드 위치 기반 분리  (fallback)
    3개 이상이면 색상·사이즈 우선 2개 선택.
    """
    if not raw_name:
        return []
    # 1) 개행으로 구분된 경우 — 운영 파일 기준
    if "\n" in raw_name:
        parts = [p.strip() for p in raw_name.split("\n") if p.strip()]
        if len(parts) >= 3:
            priority = [k for k in ["색상", "사이즈"] if k in parts]
            others = [k for k in parts if k not in priority]
            parts = (priority + others)[:2]
        return parts[:2]
    # 2) 붙은 형태 fallback — 키워드 위치 기반
    found_with_pos = []
    for kw in OPTION_KEYWORDS:
        pos = raw_name.find(kw)
        if pos != -1:
            found_with_pos.append((pos, kw))
    found_with_pos.sort(key=lambda x: x[0])
    found = [kw for _, kw in found_with_pos]
    if len(found) >= 3:
        priority = [k for k in ["색상", "사이즈"] if k in found]
        others = [k for k in found if k not in priority]
        found = (priority + others)[:2]
    return found[:2]



def parse_option_values(raw_value: str, num_attrs: int) -> list[str]:
    """
    개행 구분 옵션값 파싱.
    '블루,레드\nS,M,L' → ['블루,레드', 'S,M,L']
    개행 없으면 전체를 첫 번째 속성값으로
    """
    if "\n" in raw_value:
        parts = [p.strip() for p in raw_value.split("\n") if p.strip()]
    else:
        parts = [raw_value.strip()] if raw_value.strip() else []
    # 속성 수에 맞게 패딩
    while len(parts) < num_attrs:
        parts.append("")
    return parts[:num_attrs]


# ─────────────────────────────────────────────
# 판매가 계산
# ─────────────────────────────────────────────
def calc_sell_price(buy_price: float, margin_rate: float) -> int:
    """
    판매가 = round(매입가 × 1.1 × (1 + 마진율/100) / 10) × 10
    10원 단위 반올림
    """
    vat_price = buy_price * 1.1
    raw = vat_price * (1 + margin_rate / 100)
    return int(round(raw / 10) * 10)


# ─────────────────────────────────────────────
# 데이터 변환 핵심 로직
# ─────────────────────────────────────────────
def build_output(
    header: list[str],
    rows: list[list[str]],
    tag: str,
    margin_rate: float,
) -> list[dict]:
    """product 행 리스트를 onecell 컬럼 딕셔너리 리스트로 변환."""

    def col(row, name):
        if name in header:
            return row[header.index(name)]
        return ""

    records = []
    for row in rows:
        product_name = col(row, "상품명")
        if not product_name:
            continue

        # 상품명 태그 처리
        base_name = f"[{tag}] {product_name}" if tag else product_name

        # 판매가 계산
        try:
            buy_price = float(col(row, "판매가") or 0)
        except ValueError:
            buy_price = 0
        sell_price = calc_sell_price(buy_price, margin_rate) if buy_price > 0 else 0

        # 옵션 파싱
        raw_opt_name = col(row, "옵션명")
        raw_opt_val = col(row, "옵션값")
        opt_names = parse_option_names(raw_opt_name) if raw_opt_name else []
        opt_values = parse_option_values(raw_opt_val, len(opt_names)) if opt_names else []

        # 속성명1, 속성값1
        attr_name1 = opt_names[0] if len(opt_names) > 0 else ""
        attr_val1 = opt_values[0] if len(opt_values) > 0 else ""
        # 없으면 ONE COLOR
        if not attr_name1:
            attr_name1 = "색상"
            attr_val1 = "ONE COLOR"
        elif not attr_val1:
            attr_val1 = "ONE COLOR" if attr_name1 == "색상" else "ONE SIZE"

        # 속성명2, 속성값2
        attr_name2 = opt_names[1] if len(opt_names) > 1 else ""
        attr_val2 = opt_values[1] if len(opt_values) > 1 else ""
        if attr_name2 and not attr_val2:
            attr_val2 = "ONE SIZE" if attr_name2 == "사이즈" else ""

        records.append({
            "A": base_name,          # 기초상품명
            "B": attr_name1,         # 속성명1
            "C": attr_val1,          # 속성값1
            "D": attr_name2,         # 속성명2
            "E": attr_val2,          # 속성값2
            "H": sell_price,         # 판매가
            "J": col(row, "재고수량") or 999,  # 재고수량
            "K": col(row, "판매자 상품코드"),  # 판매자 관리 코드
            "M": col(row, "브랜드") or "상세설명참조",   # 브랜드
            "N": col(row, "제조사") or "상세설명참조",   # 제조사
            "O": "상세설명참조",      # 모델명
            "AC": col(row, "상품 상세정보"),  # 상세설명
            "AG": col(row, "상품정보제공고시 품명") or "의류",  # 상품정보제공고시 분류
            "BC": col(row, "대표 이미지 파일명"),  # 대표이미지
        })
    return records


# ─────────────────────────────────────────────
# 템플릿에 데이터 쓰기
# ─────────────────────────────────────────────
def write_to_template(records: list[dict], save_path: str):
    """onecell_template.xlsx를 열어 8행부터 데이터를 쓰고 저장."""
    from openpyxl.utils import column_index_from_string

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["기초상품정보"]

    START_ROW = 8
    for i, rec in enumerate(records):
        r = START_ROW + i
        for col_letter, value in rec.items():
            col_idx = column_index_from_string(col_letter)
            ws.cell(row=r, column=col_idx, value=value)

    wb.save(save_path)


# ─────────────────────────────────────────────
# Mock: 업로드 클래스 (추후 자동화용)
# ─────────────────────────────────────────────
class OnecellUploader:
    """
    onecell 파일 업로드 Mock 클래스.
    추후 실제 업로드 로직으로 교체 예정.

    Usage:
        uploader = OnecellUploader(credentials={"id": "...", "pw": "..."})
        result = uploader.upload(filepath="output.xlsx")
    """

    def __init__(self, credentials: dict | None = None):
        self.credentials = credentials or {}
        self._logged_in = False

    def login(self) -> bool:
        """Mock 로그인. 항상 성공 반환."""
        print(f"[Mock] 로그인 시도: {self.credentials.get('id', 'unknown')}")
        self._logged_in = True
        return True

    def upload(self, filepath: str) -> dict:
        """
        Mock 업로드.
        Returns:
            {"success": bool, "message": str, "uploaded_count": int}
        """
        if not self._logged_in:
            self.login()
        print(f"[Mock] 업로드 요청: {filepath}")
        return {
            "success": True,
            "message": f"[Mock] {os.path.basename(filepath)} 업로드 완료 (실제 업로드 미구현)",
            "uploaded_count": 0,
        }

    def logout(self):
        self._logged_in = False
        print("[Mock] 로그아웃")


# ─────────────────────────────────────────────
# GUI
# ─────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("onecell 자동 입력 도구")
        self.resizable(False, False)
        self.geometry("540x340")

        self.cfg = load_settings()
        self._build_ui()

    def _build_ui(self):
        pad = {"padx": 12, "pady": 6}

        # ── 설정 프레임 ──────────────────────────
        frm_cfg = ttk.LabelFrame(self, text="설정 (settings.ini)")
        frm_cfg.pack(fill="x", **pad)

        ttk.Label(frm_cfg, text="태그:").grid(row=0, column=0, sticky="w", padx=8, pady=4)
        self.var_tag = tk.StringVar(value=self.cfg.get("general", "tag", fallback=""))
        ttk.Entry(frm_cfg, textvariable=self.var_tag, width=20).grid(row=0, column=1, sticky="w", padx=4)
        ttk.Label(frm_cfg, text="예) 26신상  →  [26신상] 상품명").grid(row=0, column=2, sticky="w", padx=4)

        ttk.Label(frm_cfg, text="마진율 (%):").grid(row=1, column=0, sticky="w", padx=8, pady=4)
        self.var_margin = tk.StringVar(value=self.cfg.get("pricing", "margin_rate", fallback="15"))
        ttk.Entry(frm_cfg, textvariable=self.var_margin, width=8).grid(row=1, column=1, sticky="w", padx=4)
        ttk.Label(frm_cfg, text="기본값 15 | 판매가 = 매입가×1.1×(1+마진율%)").grid(row=1, column=2, sticky="w", padx=4)

        btn_save_cfg = ttk.Button(frm_cfg, text="설정 저장", command=self._save_settings)
        btn_save_cfg.grid(row=2, column=1, sticky="w", padx=4, pady=6)

        # ── 파일 선택 프레임 ─────────────────────
        frm_file = ttk.LabelFrame(self, text="상품 파일 선택")
        frm_file.pack(fill="x", **pad)

        self.var_filepath = tk.StringVar(value="")
        ttk.Entry(frm_file, textvariable=self.var_filepath, width=46, state="readonly").grid(
            row=0, column=0, padx=8, pady=8
        )
        ttk.Button(frm_file, text="파일 선택...", command=self._browse_file).grid(
            row=0, column=1, padx=4
        )

        # ── 실행 버튼 ────────────────────────────
        frm_btn = ttk.Frame(self)
        frm_btn.pack(fill="x", padx=12, pady=10)

        self.btn_run = ttk.Button(
            frm_btn, text="▶  자동 입력", command=self._run, style="Accent.TButton"
        )
        self.btn_run.pack(side="left", ipadx=16, ipady=4)

        # ── 상태 표시 ────────────────────────────
        self.var_status = tk.StringVar(value="파일을 선택하고 자동 입력을 누르세요.")
        ttk.Label(self, textvariable=self.var_status, foreground="gray").pack(
            anchor="w", padx=14, pady=(0, 8)
        )

        # ── 진행 바 ──────────────────────────────
        self.progress = ttk.Progressbar(self, mode="indeterminate", length=510)
        self.progress.pack(padx=12, pady=(0, 8))

    # ── 콜백 ────────────────────────────────────
    def _save_settings(self):
        self.cfg["general"]["tag"] = self.var_tag.get().strip()
        self.cfg["pricing"]["margin_rate"] = self.var_margin.get().strip()
        save_settings(self.cfg)
        self._status("설정이 저장되었습니다.", "green")

    def _browse_file(self):
        path = filedialog.askopenfilename(
            title="상품 파일 선택",
            filetypes=[("Excel/XLS 파일", "*.xls *.xlsx"), ("모든 파일", "*.*")],
        )
        if path:
            self.var_filepath.set(path)
            self._status(f"선택됨: {os.path.basename(path)}", "gray")

    def _run(self):
        filepath = self.var_filepath.get()
        if not filepath:
            messagebox.showwarning("파일 없음", "상품 파일을 먼저 선택해 주세요.")
            return

        # 마진율 검증
        try:
            margin = float(self.var_margin.get())
        except ValueError:
            messagebox.showerror("입력 오류", "마진율은 숫자로 입력해 주세요.")
            return

        tag = self.var_tag.get().strip()

        # 저장 경로 선택
        save_path = filedialog.asksaveasfilename(
            title="결과 파일 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            initialfile="onecell_output.xlsx",
        )
        if not save_path:
            return

        self.btn_run.config(state="disabled")
        self.progress.start(10)
        self._status("처리 중...", "blue")
        self.update()

        try:
            header, rows = parse_product_file(filepath)
            records = build_output(header, rows, tag, margin)
            write_to_template(records, save_path)

            self.progress.stop()
            self.btn_run.config(state="normal")
            self._status(f"완료: {len(records)}개 상품 → {os.path.basename(save_path)}", "green")
            messagebox.showinfo(
                "완료",
                f"{len(records)}개 상품 데이터가 저장되었습니다.\n\n{save_path}",
            )
        except Exception as e:
            self.progress.stop()
            self.btn_run.config(state="normal")
            self._status(f"오류: {e}", "red")
            messagebox.showerror("오류 발생", str(e))

    def _status(self, msg: str, color: str = "gray"):
        self.var_status.set(msg)
        for w in self.winfo_children():
            if isinstance(w, ttk.Label) and w.cget("textvariable") == str(self.var_status):
                w.config(foreground=color)
                break
        # tkinter variable label foreground 직접 적용
        self.update_idletasks()


if __name__ == "__main__":
    app = App()
    app.mainloop()