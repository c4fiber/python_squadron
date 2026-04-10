"""
app.py — UI + 설정 + 템플릿 쓰기
Flat & Material Design / KCC-Ganpan 폰트
"""
from __future__ import annotations

import configparser
import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, font as tkfont

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from onecell.parser import ProductRecord
from onecell.preset import REGISTRY, PresetConfig


# ─────────────────────────────────────────────
# 경로 헬퍼
# ─────────────────────────────────────────────
def _resource(relative: str) -> str:
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, relative)

TEMPLATE_PATH = _resource("onecell_template.xlsx")
FONT_PATH     = _resource("KCC-Ganpan.ttf")
ICON_PATH     = _resource("app_icon.ico")
SETTINGS_PATH = os.path.join(
    os.path.dirname(os.path.abspath(sys.argv[0])), "settings.ini"
)


# ─────────────────────────────────────────────
# Material Design 색상 팔레트
# ─────────────────────────────────────────────
C = {
    "bg":           "#F5F5F5",   # Grey 100 — 앱 배경
    "surface":      "#FFFFFF",   # 카드/패널 배경
    "primary":      "#1565C0",   # Blue 800
    "primary_var":  "#1976D2",   # Blue 700 (hover)
    "on_primary":   "#FFFFFF",
    "secondary":    "#E3F2FD",   # Blue 50 (선택 강조 배경)
    "accent":       "#FF6F00",   # Amber 900 (저장 버튼)
    "accent_var":   "#F57F17",
    "on_accent":    "#FFFFFF",
    "divider":      "#E0E0E0",   # Grey 300
    "text_hi":      "#212121",   # Grey 900
    "text_med":     "#757575",   # Grey 600
    "text_dis":     "#BDBDBD",   # Grey 400
    "success":      "#2E7D32",   # Green 800
    "error":        "#C62828",   # Red 800
    "warn":         "#E65100",   # Deep Orange 900
    "chip_bg":      "#E3F2FD",
    "chip_sel":     "#1565C0",
    "chip_sel_fg":  "#FFFFFF",
}


# ─────────────────────────────────────────────
# Settings
# ─────────────────────────────────────────────
def _default_cfg() -> dict:
    d = {"general": {"tag": ""}, "pricing": {"margin_rate": "15"}}
    for p in REGISTRY.values():
        d[p.settings_key] = {"shipping_fee": str(p.default_fee)}
    return d

def load_settings() -> configparser.ConfigParser:
    cfg = configparser.ConfigParser()
    if os.path.exists(SETTINGS_PATH):
        cfg.read(SETTINGS_PATH, encoding="utf-8")
    for section, kvs in _default_cfg().items():
        if not cfg.has_section(section):
            cfg[section] = kvs
        else:
            for k, v in kvs.items():
                if not cfg.has_option(section, k):
                    cfg.set(section, k, v)
    return cfg

def save_settings(cfg: configparser.ConfigParser) -> None:
    with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
        cfg.write(f)


# ─────────────────────────────────────────────
# 판매가 계산
# ─────────────────────────────────────────────
def calc_sell_price(buy: float, fee: float, margin: float) -> int:
    """판매가 = round(매입가 × 1.1 × (1 + 마진율/100) + 배송비) / 10 × 10
    배송비는 VAT/마진 계산 이후 별도 가산.
    """
    raw = buy * 1.1 * (1 + margin / 100) + fee
    return int(round(raw / 10) * 10)


# ─────────────────────────────────────────────
# 템플릿 쓰기
# ─────────────────────────────────────────────
def _write_chunk(records: list[ProductRecord], tag: str, margin: float, fee: float, save_path: str) -> None:
    """단일 청크를 템플릿에 써서 저장."""""
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["기초상품정보"]
    for i, rec in enumerate(records):
        r = 8 + i
        sp   = calc_sell_price(rec.buy_price, fee, margin) if rec.buy_price > 0 else 0
        name = f"[{tag}] {rec.product_name}" if tag else rec.product_name
        for col, val in {
            "A": name,           "B": rec.attr_name1,  "C": rec.attr_val1,
            "D": rec.attr_name2, "E": rec.attr_val2,   "H": sp,
            "J": rec.stock,      "K": rec.seller_code, "M": rec.brand,
            "N": rec.manufacturer, "O": "상세설명참조",
            "P": "N",            "Q": "과세",          "R": "Y",
            "S": "수입산",       "T": "아시아",         "U": "중국",
            "AC": rec.detail_html, "AG": rec.notice_category, "BC": rec.image_url,
        }.items():
            ws.cell(row=r, column=column_index_from_string(col), value=val)
    wb.save(save_path)


def write_to_template(
    records: list[ProductRecord],
    tag: str, margin: float, fee: float,
    save_dir: str, preset_label: str, max_records: int,
) -> list[str]:
    """
    레코드를 max_records 단위로 분할해 save_dir에 저장.
    반환값: 저장된 파일 경로 리스트.
    """
    from datetime import datetime
    records = sorted(records, key=lambda r: r.product_name.split()[-1] if r.product_name.split() else "")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    saved = []
    chunks = [records[i:i + max_records] for i in range(0, len(records), max_records)]
    for idx, chunk in enumerate(chunks):
        suffix = f"_{idx + 1}" if len(chunks) > 1 else ""
        filename = f"{preset_label}_{timestamp}{suffix}.xlsx"
        path = os.path.join(save_dir, filename)
        _write_chunk(chunk, tag, margin, fee, path)
        saved.append(path)
    return saved


# ─────────────────────────────────────────────
# Mock: 업로드
# ─────────────────────────────────────────────
class OnecellUploader:
    def __init__(self, credentials=None):
        self.credentials = credentials or {}
        self._logged_in  = False
    def login(self):
        self._logged_in = True; return True
    def upload(self, filepath):
        if not self._logged_in: self.login()
        return {"success": True, "message": f"[Mock] {os.path.basename(filepath)}", "uploaded_count": 0}
    def logout(self):
        self._logged_in = False


# ─────────────────────────────────────────────
# Material 위젯 헬퍼
# ─────────────────────────────────────────────
class _FlatButton(tk.Canvas):
    """배경색 + hover 효과를 가진 플랫 버튼."""
    def __init__(self, parent, text, command,
                 bg=None, fg=None, hover=None,
                 font=None, width=120, height=36, radius=6, **kw):
        self._bg    = bg    or C["primary"]
        self._fg    = fg    or C["on_primary"]
        self._hover = hover or C["primary_var"]
        self._cmd   = command
        self._r     = radius
        super().__init__(parent, width=width, height=height,
                         bg=parent["bg"] if hasattr(parent,"__getitem__") else C["bg"],
                         highlightthickness=0, cursor="hand2", **kw)
        self._draw(self._bg)
        self._tid = self.create_text(width//2, height//2, text=text,
                                     fill=self._fg, font=font or ("",10),
                                     anchor="center")
        self.bind("<Enter>",    lambda e: self._draw(self._hover))
        self.bind("<Leave>",    lambda e: self._draw(self._bg))
        self.bind("<Button-1>", lambda e: self._cmd())

    def _draw(self, color):
        self.delete("bg")
        w, h, r = int(self["width"]), int(self["height"]), self._r
        self.create_rectangle(r, 0, w-r, h, fill=color, outline=color, tags="bg")
        self.create_rectangle(0, r, w, h-r, fill=color, outline=color, tags="bg")
        self.create_oval(0,   0,   r*2, r*2, fill=color, outline=color, tags="bg")
        self.create_oval(w-r*2, 0,   w,   r*2, fill=color, outline=color, tags="bg")
        self.create_oval(0,   h-r*2, r*2, h,   fill=color, outline=color, tags="bg")
        self.create_oval(w-r*2, h-r*2, w,   h,   fill=color, outline=color, tags="bg")
        self.tag_raise("bg")
        if hasattr(self, "_tid"):
            self.tag_raise(self._tid)


class _Card(tk.Frame):
    """Material 카드 (흰 배경 + 하단 구분선)."""
    def __init__(self, parent, title="", **kw):
        super().__init__(parent, bg=C["surface"], **kw)
        if title:
            tk.Label(self, text=title, bg=C["surface"],
                     fg=C["primary"], font=("_ganpan", 9, "bold"),
                     anchor="w").pack(fill="x", padx=16, pady=(12,4))
            tk.Frame(self, bg=C["divider"], height=1).pack(fill="x", padx=0)


class _ChipGroup(tk.Frame):
    """Material Chip 라디오 그룹."""
    def __init__(self, parent, options: list[str], variable: tk.StringVar,
                 on_change=None, font=None, **kw):
        super().__init__(parent, bg=C["surface"], **kw)
        self._var     = variable
        self._chips   = {}
        self._on_change = on_change
        for opt in options:
            c = tk.Label(self, text=opt, cursor="hand2",
                         bg=C["chip_bg"], fg=C["primary"],
                         font=font or ("_ganpan", 9),
                         padx=14, pady=5,
                         relief="flat", bd=0)
            c.pack(side="left", padx=4)
            self._chips[opt] = c
            c.bind("<Button-1>", lambda e, o=opt: self._select(o))
        self._refresh()

    def _select(self, opt):
        self._var.set(opt)
        self._refresh()
        if self._on_change:
            self._on_change()

    def _refresh(self):
        sel = self._var.get()
        for opt, c in self._chips.items():
            if opt == sel:
                c.config(bg=C["chip_sel"], fg=C["chip_sel_fg"])
            else:
                c.config(bg=C["chip_bg"], fg=C["primary"])


class _Field(tk.Frame):
    """라벨 + 입력 필드 한 쌍."""
    def __init__(self, parent, label, variable, width=10, readonly=False, **kw):
        super().__init__(parent, bg=C["surface"], **kw)
        tk.Label(self, text=label, bg=C["surface"],
                 fg=C["text_med"], font=("_ganpan", 8),
                 anchor="w").pack(anchor="w")
        state = "readonly" if readonly else "normal"
        e = tk.Entry(self, textvariable=variable, width=width,
                     state=state,
                     font=("_ganpan", 10),
                     bg=C["bg"] if readonly else C["surface"],
                     fg=C["text_hi"],
                     insertbackground=C["primary"],
                     relief="flat", bd=0,
                     highlightthickness=1,
                     highlightbackground=C["divider"],
                     highlightcolor=C["primary"])
        e.pack(fill="x", ipady=5)
        self._entry = e


# ─────────────────────────────────────────────
# 메인 앱
# ─────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("onecell 자동 입력")
        self.resizable(False, False)
        self.configure(bg=C["bg"])
        self.cfg = load_settings()
        self._fee_vars: dict[str, tk.StringVar] = {}
        self._load_font()
        self._set_icon()
        self._build_ui()

    # ── 폰트 & 아이콘 ────────────────────────
    def _load_font(self) -> None:
        try:
            from ctypes import windll
            windll.gdi32.AddFontResourceExW(FONT_PATH, 0x10, 0)
            self._font_name = "KCC-Ganpan"
        except Exception:
            # Windows 외 환경 또는 실패 시 기본 폰트
            self._font_name = "TkDefaultFont"
        # tkinter named font 등록
        try:
            tkfont.Font(self, family=self._font_name, name="_ganpan", size=10)
            tkfont.Font(self, family=self._font_name, name="_ganpan_h", size=13, weight="bold")
            tkfont.Font(self, family=self._font_name, name="_ganpan_sm", size=8)
        except Exception:
            pass

    def _set_icon(self) -> None:
        try:
            self.iconbitmap(ICON_PATH)
        except Exception:
            pass

    # ── UI 구성 ──────────────────────────────
    def _build_ui(self) -> None:
        FN = "_ganpan"
        FN_H = "_ganpan_h"

        # 헤더 바
        header = tk.Frame(self, bg=C["primary"], height=52)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="  onecell  자동 입력 도구",
                 bg=C["primary"], fg=C["on_primary"],
                 font=(FN_H, 13, "bold"), anchor="w").pack(side="left", padx=16, fill="y")

        # 본문 패딩 프레임
        body = tk.Frame(self, bg=C["bg"])
        body.pack(fill="both", padx=16, pady=12)

        # ── 카드 1: 설정 ─────────────────────
        card1 = _Card(body, title="설정")
        card1.pack(fill="x", pady=(0, 10))
        inner1 = tk.Frame(card1, bg=C["surface"])
        inner1.pack(fill="x", padx=16, pady=12)

        # 태그 + 마진율 (가로 배치)
        row1 = tk.Frame(inner1, bg=C["surface"])
        row1.pack(fill="x", pady=(0, 10))
        self.var_tag = tk.StringVar(value=self.cfg.get("general","tag",fallback=""))
        _Field(row1, "태그", self.var_tag, width=14).pack(side="left", padx=(0,16))
        self.var_margin = tk.StringVar(value=self.cfg.get("pricing","margin_rate",fallback="15"))
        _Field(row1, "마진율 (%)", self.var_margin, width=8).pack(side="left", padx=(0,16))
        tk.Label(row1, text="판매가 = (매입가 × 1.1) × (1 + 마진율%) + 배송비",
                 bg=C["surface"], fg=C["text_med"],
                 font=(FN, 8)).pack(side="left", padx=(0,0), anchor="s", pady=4)

        # 배송비 (프리셋별 동적 생성)
        row2 = tk.Frame(inner1, bg=C["surface"])
        row2.pack(fill="x", pady=(0, 10))
        for preset in REGISTRY.values():
            var = tk.StringVar(value=self.cfg.get(
                preset.settings_key, "shipping_fee", fallback=str(preset.default_fee)))
            self._fee_vars[preset.label] = var
            _Field(row2, f"{preset.label} 배송비 (원)", var, width=10).pack(side="left", padx=(0,16))

        # 저장 버튼
        _FlatButton(inner1, "설정 저장", self._save_settings,
                    bg=C["primary"], hover=C["primary_var"],
                    font=(FN, 9), width=90, height=32).pack(anchor="w")

        # ── 카드 2: 데이터 종류 ──────────────
        card2 = _Card(body, title="데이터 종류")
        card2.pack(fill="x", pady=(0, 10))
        inner2 = tk.Frame(card2, bg=C["surface"])
        inner2.pack(fill="x", padx=16, pady=12)

        self.var_preset = tk.StringVar(value=next(iter(REGISTRY)))
        chip_row = tk.Frame(inner2, bg=C["surface"])
        chip_row.pack(side="left")
        self._chips = _ChipGroup(chip_row, list(REGISTRY.keys()),
                                  self.var_preset,
                                  on_change=self._on_preset_change,
                                  font=(FN, 9))
        self._chips.pack()

        self.lbl_fee_info = tk.Label(inner2, text="",
                                      bg=C["surface"], fg=C["text_med"],
                                      font=(FN, 8))
        self.lbl_fee_info.pack(side="left", padx=16)
        self._on_preset_change()

        # ── 카드 3: 파일 선택 ────────────────
        card3 = _Card(body, title="상품 파일")
        card3.pack(fill="x", pady=(0, 10))
        inner3 = tk.Frame(card3, bg=C["surface"])
        inner3.pack(fill="x", padx=16, pady=12)

        self.var_filepath = tk.StringVar()
        file_row = tk.Frame(inner3, bg=C["surface"])
        file_row.pack(fill="x")
        _Field(file_row, "파일 경로", self.var_filepath,
               width=46, readonly=True).pack(side="left", fill="x", expand=True, padx=(0,10))
        _FlatButton(file_row, "찾아보기", self._browse_file,
                    bg=C["primary"], hover=C["primary_var"],
                    font=(FN, 9), width=80, height=32).pack(side="left", anchor="s", pady=4)

        # ── 실행 버튼 ────────────────────────
        action = tk.Frame(body, bg=C["bg"])
        action.pack(fill="x", pady=(4, 0))
        self.btn_run = _FlatButton(action, "▶  자동 입력", self._run,
                                    bg=C["accent"], hover=C["accent_var"],
                                    fg=C["on_accent"],
                                    font=(FN, 10, "bold"),
                                    width=140, height=40)
        self.btn_run.pack(side="left")

        # ── 상태 바 ──────────────────────────
        status_bar = tk.Frame(self, bg=C["divider"], height=1)
        status_bar.pack(fill="x")
        self.var_status = tk.StringVar(value="파일을 선택하고 자동 입력을 누르세요.")
        self.lbl_status = tk.Label(self, textvariable=self.var_status,
                                    bg=C["bg"], fg=C["text_med"],
                                    font=(FN, 8), anchor="w")
        self.lbl_status.pack(fill="x", padx=16, pady=(6, 2))

        # 진행 바 (Canvas 기반 플랫)
        self._prog_canvas = tk.Canvas(self, height=4, bg=C["divider"],
                                       highlightthickness=0)
        self._prog_canvas.pack(fill="x", padx=0, pady=(0, 8))
        self._prog_rect = None
        self._prog_anim = False

    # ── 진행 바 애니메이션 ────────────────────
    def _prog_start(self):
        self._prog_anim = True
        self._prog_pos  = 0.0
        self._prog_step()

    def _prog_step(self):
        if not self._prog_anim:
            return
        w = self._prog_canvas.winfo_width() or 580
        self._prog_canvas.delete("bar")
        bar_w = w * 0.35
        x1 = (self._prog_pos % 1.35) * w - bar_w * 0.2
        x2 = x1 + bar_w
        self._prog_canvas.create_rectangle(
            max(0, x1), 0, min(w, x2), 4,
            fill=C["primary"], outline="", tags="bar")
        self._prog_pos += 0.018
        self.after(16, self._prog_step)

    def _prog_stop(self):
        self._prog_anim = False
        self._prog_canvas.delete("bar")

    # ── 콜백 ─────────────────────────────────
    def _on_preset_change(self) -> None:
        label = self.var_preset.get()
        var   = self._fee_vars.get(label)
        try:
            fee = int(float(var.get())) if var else 0
            self.lbl_fee_info.config(text=f"배송비  {fee:,}원")
        except ValueError:
            self.lbl_fee_info.config(text="")

    def _save_settings(self) -> None:
        self.cfg["general"]["tag"]         = self.var_tag.get().strip()
        self.cfg["pricing"]["margin_rate"] = self.var_margin.get().strip()
        for preset in REGISTRY.values():
            var = self._fee_vars.get(preset.label)
            if var:
                self.cfg[preset.settings_key]["shipping_fee"] = var.get().strip()
        save_settings(self.cfg)
        self._set_status("설정이 저장되었습니다.", C["success"])

    def _browse_file(self) -> None:
        preset = REGISTRY[self.var_preset.get()]
        path = filedialog.askopenfilename(
            title="상품 파일 선택", filetypes=preset.file_types)
        if path:
            self.var_filepath.set(path)
            self._set_status(f"선택됨: {os.path.basename(path)}", C["text_med"])

    def _run(self) -> None:
        filepath = self.var_filepath.get()
        if not filepath:
            messagebox.showwarning("파일 없음", "상품 파일을 먼저 선택해 주세요.")
            return
        try:
            margin = float(self.var_margin.get())
        except ValueError:
            messagebox.showerror("입력 오류", "마진율은 숫자로 입력해 주세요.")
            return
        label  = self.var_preset.get()
        preset = REGISTRY[label]
        tag    = self.var_tag.get().strip()
        try:
            fee = float(self._fee_vars[label].get())
        except ValueError:
            messagebox.showerror("입력 오류", "배송비는 숫자로 입력해 주세요.")
            return
        save_dir = filedialog.askdirectory(title="결과 파일 저장 폴더 선택")
        if not save_dir:
            return

        self.btn_run.config(state="disabled")
        self._prog_start()
        self._set_status("처리 중...", C["primary"])
        self.update()

        try:
            records = preset.make_parser().parse(filepath)
            saved = write_to_template(
                records, tag, margin, fee,
                save_dir, label, preset.max_records)
            self._prog_stop()
            self.btn_run.config(state="normal")
            files_str = "\n".join(os.path.basename(p) for p in saved)
            self._set_status(
                f"완료  {len(records)}개 상품  →  {len(saved)}개 파일",
                C["success"])
            messagebox.showinfo(
                "완료",
                f"{len(records)}개 상품을 {len(saved)}개 파일로 저장했습니다.\n\n{files_str}\n\n저장 위치: {save_dir}")
        except Exception as exc:
            self._prog_stop()
            self.btn_run.config(state="normal")
            self._set_status(f"오류: {exc}", C["error"])
            messagebox.showerror("오류 발생", str(exc))

    def _set_status(self, msg: str, color: str = "") -> None:
        self.var_status.set(msg)
        self.lbl_status.config(foreground=color or C["text_med"])
        self.update_idletasks()